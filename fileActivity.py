from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import List
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import uvicorn
app = FastAPI()


class Activity(BaseModel):
    name: str
    location: str = None
    duration_minutes: int
    
FILE_PATH = "activities.xlsx"

def initialize_excel():
    if not os.path.exists(FILE_PATH):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["ID", "Name", "Location", "Duration"])
        workbook.save(FILE_PATH)

def get_activities_from_excel():
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    activities = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        activities.append({
            "id": row[0],
            "name": row[1],
            "location": row[2],
            "duration_minutes": row[3],
        })
    return activities

def write_activity_to_excel(activity: Activity):
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    next_id = sheet.max_row
    sheet.append([next_id, activity.name, activity.location, activity.duration_minutes])
    workbook.save(FILE_PATH)
    return next_id

def update_activity_in_excel(activity_id: int, updated_activity: Activity):
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == activity_id:
            row[1].value = updated_activity.name
            row[2].value = updated_activity.location
            row[3].value = updated_activity.duration_minutes
            workbook.save(FILE_PATH)
            return
    raise HTTPException(status_code=404, detail="Activity not found")


def delete_activity_from_excel(activity_id: int):
    workbook = load_workbook(FILE_PATH)
    sheet = workbook.active
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[0].value == activity_id:
            sheet.delete_rows(row[0].row, 1)
            workbook.save(FILE_PATH)
            return
    raise HTTPException(status_code=404, detail="Activity not found")


initialize_excel()

@app.get("/", response_model=List[Activity])
def get_activities():
    activities = get_activities_from_excel()
    return activities

@app.post("/activities/", response_model=Activity)
def create_activity(activity: Activity):
    activity_id = write_activity_to_excel(activity)
    return {**activity.model_dump(), "id": activity_id}

@app.put("/activities/{activity_id}", response_model=Activity)
def update_activity(activity_id: int, updated_activity: Activity):
    update_activity_in_excel(activity_id, updated_activity)
    return {**updated_activity.model_dump(), "id": activity_id}


@app.delete("/activities/{activity_id}", response_model=Activity)
def delete_activity(activity_id: int):
    activities = get_activities_from_excel()
    activity = next((act for act in activities if act["id"] == activity_id), None)
    if not activity:
        raise HTTPException(status_code=404, detail="Activity not found")
    delete_activity_from_excel(activity_id)
    return activity

if __name__ == "__main__":
    uvicorn.run(app, host="127.0.0.1", port=5000, log_level="info")